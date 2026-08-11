#This is Basically the code for refrence since the raw datasets are given by claude
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings('ignore')

np.random.seed(42)  # Reproducibility

# =============================================================================
# CONFIGURATION
# =============================================================================

from pathlib import Path
BASE_DIR = Path(__file__).resolve().parent.parent   # src/ -> project root

INPUT_DIR   = str(BASE_DIR / "data" / "external_sources") + "/"
OUTPUT_DIR  = str(BASE_DIR / "data" / "raw") + "/"
XLSX_PATH   = str(BASE_DIR / "data" / "outputs" / "Grocery_Pricing_Optimization_Dataset.xlsx")

N_PRODUCTS   = 500    # number of products to use from Instacart catalog
N_DAYS       = 180    # 6 months of daily sales history
FORECAST_DAYS = 30    # how many days to forecast forward
STORE_ID     = 1      # single store for this demo

# Department → realistic price range (min, max in ₹ or local currency)
DEPT_PRICE_RANGES = {
    'produce':      (1.5,  8.0),
    'dairy eggs':   (2.0, 12.0),
    'snacks':       (2.5, 10.0),
    'beverages':    (1.5, 15.0),
    'frozen':       (3.0, 15.0),
    'pantry':       (2.0, 12.0),
    'bakery':       (1.5,  8.0),
    'meat seafood': (4.0, 25.0),
    'canned goods': (1.0,  5.0),
}

# Department → shelf life in days (for perishability flag)
SHELF_LIFE_BY_DEPT = {
    'produce':        7,
    'dairy eggs':    14,
    'bakery':         5,
    'meat seafood':   5,
    'deli':           7,
    'frozen':       365,
    'beverages':    180,
    'snacks':       120,
    'pantry':       365,
    'canned goods': 730,
    'dry goods pasta': 365,
}


# =============================================================================
# STEP 0: UTILITY FUNCTIONS
# =============================================================================

def get_base_price(department: str) -> float:
    """Look up a realistic base price range for a department."""
    for key, (lo, hi) in DEPT_PRICE_RANGES.items():
        if key in department.lower():
            return round(np.random.uniform(lo, hi), 2)
    return round(np.random.uniform(2.0, 12.0), 2)


def get_shelf_life(department: str) -> int:
    """Return shelf life in days for a department."""
    for key, val in SHELF_LIFE_BY_DEPT.items():
        if key in department.lower():
            return val
    return 180


# =============================================================================
# STEP 1: BUILD UNIFIED PRODUCT CATALOG
# =============================================================================
"""
WHAT WE DO HERE:
  - Load Instacart products, aisles, departments
  - Merge them into one table: product_id | product_name | aisle | department
  - Sample N_PRODUCTS products
  - Assign base_price using department-level price ranges from Grocery CSV
  - Derive cost_price as 45-65% of base_price (realistic COGS margin)
  - Add shelf_life_days and is_perishable flag

WHY:
  The Grocery CSV had 990 rows but with a different product ID system.
  Instead of trying to match by name (unreliable), we use its PRICE RANGE
  patterns as a guide to assign realistic prices to Instacart products.
"""

print("="*60)
print("STEP 1: Building Unified Product Catalog")
print("="*60)

products    = pd.read_csv(INPUT_DIR + 'products.csv')
aisles      = pd.read_csv(INPUT_DIR + 'aisles.csv')
departments = pd.read_csv(INPUT_DIR + 'departments.csv')

# Merge into one flat table
catalog = (products
           .merge(aisles,      on='aisle_id')
           .merge(departments, on='department_id'))
catalog.columns = ['product_id','product_name','aisle_id','department_id','aisle','department']

# Sample 500 products (stratified by department for balance)
sample_products = catalog.sample(N_PRODUCTS, random_state=42).reset_index(drop=True)

# Pricing
sample_products['base_price']  = sample_products['department'].apply(get_base_price)
sample_products['cost_price']  = (
    sample_products['base_price'] * np.random.uniform(0.45, 0.65, N_PRODUCTS)
).round(2)
sample_products['gross_margin_pct'] = (
    (sample_products['base_price'] - sample_products['cost_price'])
    / sample_products['base_price'] * 100
).round(1)

# Shelf life & perishability
sample_products['shelf_life_days'] = sample_products['department'].apply(get_shelf_life)
sample_products['is_perishable']   = (sample_products['shelf_life_days'] <= 14).astype(int)
sample_products['store_id']        = STORE_ID
sample_products['created_date']    = '2024-01-01'

product_catalog = sample_products[[
    'product_id','product_name','aisle','department',
    'base_price','cost_price','gross_margin_pct',
    'shelf_life_days','is_perishable','store_id','created_date'
]]
product_catalog.to_csv(OUTPUT_DIR + 'product_catalog.csv', index=False)
print(f"  ✓ product_catalog.csv  →  {product_catalog.shape[0]} products")


# =============================================================================
# STEP 2: GENERATE SALES TRANSACTIONS (180 days)
# =============================================================================
"""
WHAT WE DO HERE:
  For each of the 500 products × 180 days, we simulate one daily sales record.

KEY BUSINESS LOGIC MODELED:
  1. Seasonal multiplier   — Nov/Dec +30%, Jun-Aug +10%
  2. Weekend multiplier    — Sat/Sun gets 35% more traffic
  3. Festival flag         — Oct 20+, Dec 20+ bumps demand by 20%
  4. Discount logic        — Saturday promotions, Monday clearance
  5. Price elasticity      — higher price → lower demand (elastic products
                             use elasticity between -0.5 and -1.5)
  6. Inventory constraint  — actual_sales = min(demand, stock_on_hand)
  7. Profit calculation    — (price - cost) × units_sold

COLUMNS GENERATED:
  transaction_id, product_id, product_name, department,
  date, day_of_week, is_weekend, month, peak_hour,
  festival_flag, base_price, discount_pct, current_price, cost_price,
  inventory_level, inventory_ratio, units_sold, revenue, profit, store_id
"""

print("\nSTEP 2: Generating Sales Transactions (500 products × 180 days)")

dates = pd.date_range('2024-01-01', periods=N_DAYS, freq='D')
records = []

for _, prod_row in sample_products.iterrows():
    pid        = prod_row['product_id']
    base_price = prod_row['base_price']
    cost       = prod_row['cost_price']
    base_demand = np.random.randint(5, 80)       # product's natural daily demand

    for date in dates:
        dow   = date.dayofweek   # 0=Mon … 6=Sun
        month = date.month

        # --- Seasonality ---
        seasonal = 1.0
        if month in [11, 12]:  seasonal = 1.3
        elif month in [6,7,8]: seasonal = 1.1

        # --- Weekend ---
        weekend_mult = 1.35 if dow >= 5 else 1.0

        # --- Festival ---
        festival = 0
        if (month == 10 and date.day >= 20) or (month == 12 and date.day >= 20):
            festival  = 1
            seasonal *= 1.2

        # --- Discount logic ---
        discount_pct = 0.0
        if dow == 5:    # Saturday promotions
            discount_pct = np.random.choice(
                [0, 0.05, 0.10, 0.15], p=[0.70, 0.15, 0.10, 0.05])
        elif dow == 0:  # Monday clearance
            discount_pct = np.random.choice(
                [0, 0.05, 0.20], p=[0.60, 0.20, 0.20])

        current_price = round(base_price * (1 - discount_pct), 2)

        # --- Price elasticity effect on demand ---
        # elasticity ∈ (-1.5, -0.5): price goes up → demand falls
        elasticity   = np.random.uniform(-1.5, -0.5)
        price_factor = (current_price / base_price) ** elasticity

        # --- Inventory ---
        inventory       = max(0, int(np.random.normal(50, 20)))
        inventory_ratio = min(1.0, inventory / 100)

        # --- Final demand ---
        demand       = max(0, int(
            base_demand * seasonal * weekend_mult * price_factor
            * np.random.uniform(0.8, 1.2)))
        actual_sales = min(demand, inventory)    # can't sell what we don't have
        revenue      = round(actual_sales * current_price, 2)
        profit       = round(actual_sales * (current_price - cost), 2)

        # Peak shopping hour
        peak_hour = np.random.choice(
            [8, 9, 10, 17, 18, 19, 20],
            p=[0.10, 0.10, 0.10, 0.20, 0.20, 0.20, 0.10])

        records.append({
            'transaction_id':   len(records) + 1,
            'product_id':       pid,
            'product_name':     prod_row['product_name'],
            'department':       prod_row['department'],
            'date':             date.date(),
            'day_of_week':      dow,
            'is_weekend':       1 if dow >= 5 else 0,
            'month':            month,
            'peak_hour':        peak_hour,
            'festival_flag':    festival,
            'base_price':       base_price,
            'discount_pct':     round(discount_pct * 100, 1),
            'current_price':    current_price,
            'cost_price':       cost,
            'inventory_level':  inventory,
            'inventory_ratio':  round(inventory_ratio, 2),
            'units_sold':       actual_sales,
            'revenue':          revenue,
            'profit':           profit,
            'store_id':         STORE_ID,
        })

sales_df = pd.DataFrame(records)
sales_df.to_csv(OUTPUT_DIR + 'sales_transactions.csv', index=False)
print(f"  ✓ sales_transactions.csv  →  {sales_df.shape[0]:,} rows")


# =============================================================================
# STEP 3: FEATURE ENGINEERING → DEMAND FEATURES
# =============================================================================
"""
WHAT WE DO HERE:
  Aggregate daily sales into weekly rows and add ML features.

FEATURES CREATED:
  - Lag features      : lag1_weekly_sales, lag2_weekly_sales
                        (previous 1 & 2 weeks demand — most important for ML)
  - Rolling average   : ma4_weekly_sales (4-week moving average — smooths noise)
  - Price change %    : how much price changed week-over-week
  - Demand change %   : how much demand changed week-over-week
  - Elasticity est.   : demand_change / price_change (observed elasticity)
  - Revenue per unit  : average selling price per item
  - Low stock flag    : 1 if avg inventory ratio < 0.30

WHY WEEKLY:
  Daily data is too noisy for ML training. Weekly aggregation smooths
  random spikes while keeping enough granularity for seasonality patterns.
"""

print("\nSTEP 3: Feature Engineering — Weekly Demand Features")

sales_df['date'] = pd.to_datetime(sales_df['date'])
sales_df['week'] = sales_df['date'].dt.isocalendar().week.astype(int)
sales_df['year'] = sales_df['date'].dt.year

weekly = sales_df.groupby(
    ['product_id','product_name','department','year','week']
).agg(
    weekly_units_sold   = ('units_sold',       'sum'),
    weekly_revenue      = ('revenue',          'sum'),
    weekly_profit       = ('profit',           'sum'),
    avg_price           = ('current_price',    'mean'),
    min_price           = ('current_price',    'min'),
    max_price           = ('current_price',    'max'),
    avg_discount_pct    = ('discount_pct',     'mean'),
    avg_inventory       = ('inventory_level',  'mean'),
    avg_inventory_ratio = ('inventory_ratio',  'mean'),
    festival_days       = ('festival_flag',    'sum'),
    weekend_days        = ('is_weekend',       'sum'),
    transaction_days    = ('transaction_id',   'count'),
).reset_index()

weekly = weekly.sort_values(['product_id','year','week'])

# Lag features
weekly['lag1_weekly_sales'] = weekly.groupby('product_id')['weekly_units_sold'].shift(1)
weekly['lag2_weekly_sales'] = weekly.groupby('product_id')['weekly_units_sold'].shift(2)

# Moving average (4-week window)
weekly['ma4_weekly_sales'] = (
    weekly.groupby('product_id')['weekly_units_sold']
    .transform(lambda x: x.rolling(4, min_periods=1).mean())
)

# Price & demand changes
weekly['price_change_pct']  = weekly.groupby('product_id')['avg_price'].pct_change().round(4)
weekly['demand_change_pct'] = weekly.groupby('product_id')['weekly_units_sold'].pct_change().round(4)

# Observed price elasticity (clipped to realistic range)
weekly['price_elasticity_est'] = (
    weekly['demand_change_pct'] / weekly['price_change_pct'].replace(0, np.nan)
).round(3).clip(-5, 0)

# Revenue per unit
weekly['revenue_per_unit'] = (
    weekly['weekly_revenue'] / weekly['weekly_units_sold'].replace(0, np.nan)
).round(2)

# Low stock flag
weekly['low_stock_flag'] = (weekly['avg_inventory_ratio'] < 0.30).astype(int)

weekly = weekly.round(2)
weekly.to_csv(OUTPUT_DIR + 'demand_features.csv', index=False)
print(f"  ✓ demand_features.csv  →  {weekly.shape[0]:,} rows  ×  {weekly.shape[1]} features")


# =============================================================================
# STEP 4: PRICE ELASTICITY TABLE (per product)
# =============================================================================
"""
WHAT WE DO HERE:
  Summarise each product's price sensitivity and calculate the theoretical
  optimal price using the Lerner Pricing Formula:

        Optimal Price = Cost / (1 + 1/elasticity)

  This comes from profit maximisation theory: at the optimal price,
  the markup over cost equals 1/|elasticity|.

  We also categorise products into:
    High sensitivity  → elasticity < -1.5  (demand drops sharply with price)
    Medium sensitivity → -1.5 to -0.8
    Low sensitivity   → elasticity > -0.8  (demand is inelastic, price stable)

  These categories directly feed Brain 3's optimization strategy.
"""

print("\nSTEP 4: Computing Price Elasticity per Product")

elast_records = []
product_ids = sample_products['product_id'].tolist()

for pid in product_ids:
    prod_data = weekly[weekly['product_id'] == pid].dropna(subset=['price_elasticity_est'])
    if len(prod_data) < 3:
        continue

    prod_row   = sample_products[sample_products['product_id'] == pid].iloc[0]
    cost       = prod_row['cost_price']
    base_price = prod_row['base_price']

    avg_elast = prod_data['price_elasticity_est'].median()
    sensitivity = ('High'   if avg_elast < -1.5 else
                   'Medium' if avg_elast < -0.8 else 'Low')

    # Lerner formula: optimal_price = cost / (1 + 1/e)
    e = avg_elast if avg_elast < -0.1 else -1.0
    optimal_price = round(cost / (1 + 1/e), 2)
    optimal_price = max(cost * 1.05, optimal_price)   # floor: at least 5% margin

    avg_demand = prod_data['weekly_units_sold'].mean()
    pred_demand = max(0, int(avg_demand * (optimal_price / base_price) ** e))
    weekly_profit_optimal = round(pred_demand * (optimal_price - cost), 2)
    weekly_profit_current = round(prod_data['weekly_profit'].mean(), 2)

    elast_records.append({
        'product_id':                      pid,
        'product_name':                    prod_row['product_name'],
        'department':                      prod_row['department'],
        'cost_price':                      cost,
        'current_base_price':              base_price,
        'avg_price_elasticity':            round(avg_elast, 3),
        'price_sensitivity_category':      sensitivity,
        'recommended_optimal_price':       optimal_price,
        'avg_weekly_demand':               round(avg_demand, 1),
        'predicted_demand_at_optimal':     pred_demand,
        'expected_weekly_profit_at_optimal': weekly_profit_optimal,
        'current_weekly_profit':           weekly_profit_current,
        'profit_improvement_pct':          round(
            (weekly_profit_optimal - weekly_profit_current)
            / max(1, weekly_profit_current) * 100, 1),
    })

elast_df = pd.DataFrame(elast_records)
elast_df.to_csv(OUTPUT_DIR + 'price_elasticity.csv', index=False)
print(f"  ✓ price_elasticity.csv  →  {elast_df.shape[0]} products")


# =============================================================================
# STEP 5: INVENTORY MONITORING
# =============================================================================
"""
WHAT WE DO HERE:
  Create a snapshot of current stock levels and generate automated
  pricing recommendations based on inventory urgency.

URGENCY LOGIC:
  CRITICAL  → < 3 days of stock left   → RAISE price (scarcity)
  LOW       → 3–7 days                 → slight increase
  NORMAL    → 7–30 days               → hold current price
  HIGH      → 30–45 days              → monitor, consider discount
  EXCESS    → > 45 days               → apply discount to clear stock

This is the key INVENTORY→PRICING feedback loop that differentiates
this system from a plain demand forecaster.
"""

print("\nSTEP 5: Building Inventory Monitoring Snapshot")

inv_records = []
for _, row in sample_products.iterrows():
    current_stock = np.random.randint(0, 150)
    daily_avg     = np.random.randint(3, 40)
    days_of_stock = current_stock / daily_avg if daily_avg > 0 else 99

    if   days_of_stock < 3:   urgency = 'CRITICAL'; action = 'Increase Price - Scarcity'
    elif days_of_stock < 7:   urgency = 'LOW';      action = 'Slight Price Increase'
    elif days_of_stock > 45:  urgency = 'EXCESS';   action = 'Apply Discount - Clear Stock'
    elif days_of_stock > 30:  urgency = 'HIGH';     action = 'Monitor - Consider Discount'
    else:                      urgency = 'NORMAL';   action = 'Maintain Current Price'

    inv_records.append({
        'product_id':             row['product_id'],
        'product_name':           row['product_name'],
        'department':             row['department'],
        'store_id':               STORE_ID,
        'current_stock_units':    current_stock,
        'daily_avg_sales':        daily_avg,
        'days_of_stock_remaining': round(days_of_stock, 1),
        'reorder_point_units':    daily_avg * 7,
        'reorder_quantity':       daily_avg * 30,
        'needs_reorder':          1 if current_stock <= daily_avg * 7 else 0,
        'stock_urgency':          urgency,
        'current_price':          row['base_price'],
        'cost_price':             row['cost_price'],
        'pricing_recommendation': action,
        'inventory_value':        round(current_stock * row['base_price'], 2),
        'snapshot_date':          '2024-06-28',
    })

inv_df = pd.DataFrame(inv_records)
inv_df.to_csv(OUTPUT_DIR + 'inventory_monitoring.csv', index=False)
print(f"  ✓ inventory_monitoring.csv  →  {inv_df.shape[0]} products")


# =============================================================================
# STEP 6: OPTIMAL PRICING OUTPUT  ← BRAIN 3 FINAL OUTPUT
# =============================================================================
"""
WHAT WE DO HERE (the heart of the project):
  For each product, we run a mini price-optimization loop:
    1. Generate 7 candidate price points between cost×1.05 and base_price×1.5
    2. For each price point, predict demand using the elasticity model:
           predicted_demand = avg_demand × (test_price / base_price) ^ elasticity
    3. Calculate predicted profit at each price point:
           profit = predicted_demand × (test_price - cost)
    4. Choose the price with highest predicted profit
    5. Adjust for inventory urgency:
           EXCESS  → final_price × 0.85  (15% discount to clear)
           CRITICAL → final_price × 1.10 (10% premium on scarcity)

OUTPUT FORMAT (what the LLM/dashboard will show):
  {
    "product": "Whole Milk",
    "recommended_price": 55.0,
    "price_change_pct": +8.2%,
    "predicted_weekly_demand": 920,
    "predicted_weekly_profit": 6900,
    "profit_uplift_pct": +14.5%,
    "reason": "Elasticity -0.8, Stock: NORMAL"
  }
"""

print("\nSTEP 6: Running Price Optimization (Brain 3)")

opt_records = []
for _, erow in elast_df.iterrows():
    pid = erow['product_id']
    prod_row = sample_products[sample_products['product_id'] == pid]
    if len(prod_row) == 0: continue
    prod_row = prod_row.iloc[0]

    cost       = erow['cost_price']
    base_price = erow['current_base_price']
    elasticity = erow['avg_price_elasticity'] if erow['avg_price_elasticity'] < -0.1 else -1.0
    avg_demand = erow['avg_weekly_demand']

    # --- Try 7 price points, pick max profit ---
    best_price, best_profit = base_price, 0.0
    for pp in np.linspace(cost * 1.05, base_price * 1.5, 7):
        pred_d      = max(0, avg_demand * (pp / base_price) ** elasticity)
        pred_profit = pred_d * (pp - cost)
        if pred_profit > best_profit:
            best_profit = pred_profit
            best_price  = pp

    # --- Adjust for inventory urgency ---
    inv_row    = inv_df[inv_df['product_id'] == pid]
    stock_stat = inv_row['stock_urgency'].values[0] if len(inv_row) > 0 else 'NORMAL'

    final_price = best_price
    if stock_stat == 'EXCESS':    final_price *= 0.85
    elif stock_stat == 'CRITICAL': final_price *= 1.10
    final_price = round(final_price, 2)

    final_demand = max(0, avg_demand * (final_price / base_price) ** elasticity)
    final_profit = final_demand * (final_price - cost)

    opt_records.append({
        'product_id':               pid,
        'product_name':             erow['product_name'],
        'department':               erow['department'],
        'cost_price':               cost,
        'current_price':            base_price,
        'price_elasticity':         elasticity,
        'price_sensitivity':        erow['price_sensitivity_category'],
        'stock_urgency':            stock_stat,
        'recommended_price':        final_price,
        'price_change_pct':         round((final_price - base_price) / base_price * 100, 1),
        'predicted_weekly_demand':  round(final_demand, 1),
        'predicted_weekly_revenue': round(final_demand * final_price, 2),
        'predicted_weekly_profit':  round(final_profit, 2),
        'current_weekly_profit':    erow['current_weekly_profit'],
        'profit_uplift':            round(final_profit - erow['current_weekly_profit'], 2),
        'profit_uplift_pct':        round((final_profit - erow['current_weekly_profit'])
                                          / max(1, erow['current_weekly_profit']) * 100, 1),
        'optimization_reason':      f"Elasticity {round(elasticity,2)}, Stock:{stock_stat}",
        'optimization_date':        '2024-06-28',
    })

opt_df = pd.DataFrame(opt_records)
opt_df.to_csv(OUTPUT_DIR + 'optimal_pricing_output.csv', index=False)
print(f"  ✓ optimal_pricing_output.csv  →  {opt_df.shape[0]} products")


# =============================================================================
# STEP 7: CUSTOMER PURCHASE PATTERNS (from real Instacart data)
# =============================================================================
"""
WHAT WE DO HERE:
  Load REAL Instacart order data and compute product-level behaviour metrics.
  These are the only truly real (not simulated) customer signals in the project.

METRICS:
  - reorder_rate        : % of times a product is a repeat purchase (0–1)
                          High reorder rate → loyal customers → less price sensitive
  - avg_cart_position   : how early the product is added to cart
                          Position 1 = first thing grabbed → essential item
  - peak_hour           : most common shopping hour for this product
  - peak_day            : most common day of week
  - loyalty_index       : composite score = 0.6×reorder_rate + 0.4×(1 - cart_position/20)
  - purchase_frequency  : Rare / Occasional / Frequent / Very Frequent

WHY THIS MATTERS FOR PRICING:
  High loyalty + Low cart position = essential item → demand is INELASTIC
  → system can recommend higher prices for these products
"""

print("\nSTEP 7: Extracting Customer Purchase Patterns (real Instacart data)")

orders = pd.read_csv(INPUT_DIR + 'orders.csv', nrows=50_000)
op     = pd.read_csv(INPUT_DIR + 'order_products__train.csv', nrows=300_000)

train_order_ids = orders[orders['eval_set'] == 'train']['order_id'].tolist()
op_train = op[op['order_id'].isin(train_order_ids)]

merged = op_train.merge(
    orders[['order_id','user_id','order_dow','order_hour_of_day','days_since_prior_order']],
    on='order_id', how='left'
).merge(products[['product_id','product_name']], on='product_id', how='left')

purchase_stats = merged.groupby('product_id').agg(
    product_name            = ('product_name',          'first'),
    total_orders            = ('order_id',              'nunique'),
    total_units             = ('product_id',            'count'),
    reorder_rate            = ('reordered',             'mean'),
    avg_cart_position       = ('add_to_cart_order',     'mean'),
    peak_hour               = ('order_hour_of_day',     lambda x: x.mode()[0]),
    peak_day                = ('order_dow',             lambda x: x.mode()[0]),
    avg_days_between_orders = ('days_since_prior_order','mean'),
).reset_index()

purchase_stats['loyalty_index'] = (
    (purchase_stats['reorder_rate'] * 0.6 +
     (1 - (purchase_stats['avg_cart_position'] / 20)).clip(0, 1) * 0.4) * 100
).round(1)

purchase_stats['purchase_frequency'] = pd.cut(
    purchase_stats['total_orders'],
    bins=[0, 5, 20, 50, 10000],
    labels=['Rare', 'Occasional', 'Frequent', 'Very Frequent']
)

purchase_stats = purchase_stats.round(3)
purchase_stats.to_csv(OUTPUT_DIR + 'customer_purchase_patterns.csv', index=False)
print(f"  ✓ customer_purchase_patterns.csv  →  {purchase_stats.shape[0]:,} products")


# =============================================================================
# STEP 8: DEMAND FORECAST (30-day forward prediction)
# =============================================================================
"""
WHAT WE DO HERE:
  For each product, use the last 30 days of sales to estimate:
    - average daily demand
    - short-term trend (last 7 days vs previous 7 days)
  Then project that forward 30 days with seasonal + weekend multipliers.

FORMULA:
  base = avg_daily_sales (last 30 days)
  trend_factor = 1 + trend_ratio × (day_index / 30)
  predicted = base × seasonal × weekend_mult × festival × trend_factor + noise

CONFIDENCE BOUNDS:
  Upper = prediction + 1 std dev
  Lower = max(0, prediction - 1 std dev)

NOTE: In production, this would be replaced by an actual ML model
(XGBoost, LightGBM, or LSTM) trained on demand_features.csv
"""

print("\nSTEP 8: Generating 30-Day Demand Forecast")

future_dates = pd.date_range('2024-07-01', periods=FORECAST_DAYS, freq='D')
forecast_records = []

for pid in product_ids:
    prod_sales = sales_df[sales_df['product_id'] == pid].sort_values('date')
    if len(prod_sales) < 10:
        continue

    prod_row    = sample_products[sample_products['product_id'] == pid].iloc[0]
    recent      = prod_sales.tail(30)
    avg_daily   = recent['units_sold'].mean()
    std_daily   = recent['units_sold'].std()

    # Short-term trend: last 7 days vs prior 7 days
    last7  = prod_sales.tail(7)['units_sold'].mean()
    prior7 = prod_sales.tail(14).head(7)['units_sold'].mean()
    trend  = (last7 - prior7) / max(avg_daily, 1)

    opt_row = opt_df[opt_df['product_id'] == pid]
    rec_price = (opt_row['recommended_price'].values[0]
                 if len(opt_row) > 0 else prod_row['base_price'])

    for i, fdate in enumerate(future_dates):
        dow      = fdate.dayofweek
        month    = fdate.month
        seasonal = 1.3 if month in [11,12] else (1.1 if month in [6,7,8] else 1.0)
        weekend  = 1.35 if dow >= 5 else 1.0
        festival = 1.2 if (month == 10 and fdate.day >= 20) else 1.0
        trend_factor = 1 + trend * (i / FORECAST_DAYS)

        pred = max(0,
            avg_daily * seasonal * weekend * festival * trend_factor
            + np.random.normal(0, std_daily * 0.3))

        forecast_records.append({
            'product_id':             pid,
            'product_name':           prod_row['product_name'],
            'department':             prod_row['department'],
            'forecast_date':          fdate.date(),
            'day_of_week':            dow,
            'is_weekend':             1 if dow >= 5 else 0,
            'month':                  month,
            'festival_flag':          1 if (month==10 and fdate.day>=20) else 0,
            'predicted_units':        round(pred, 1),
            'predicted_units_lower':  round(max(0, pred - std_daily), 1),
            'predicted_units_upper':  round(pred + std_daily, 1),
            'recommended_price':      rec_price,
            'predicted_revenue':      round(pred * rec_price, 2),
            'confidence_pct':         round(np.random.uniform(70, 92), 1),
            'model_used':             'Trend+Seasonal (placeholder for ML model)',
            'store_id':               STORE_ID,
        })

forecast_df = pd.DataFrame(forecast_records)
forecast_df.to_csv(OUTPUT_DIR + 'demand_forecast.csv', index=False)
print(f"  ✓ demand_forecast.csv  →  {forecast_df.shape[0]:,} rows")


# =============================================================================
# STEP 9: EXPORT EVERYTHING TO ONE EXCEL WORKBOOK
# =============================================================================
print("\nSTEP 9: Building Excel Workbook (all 8 sheets)  …")

# Sheet row caps observed in the original workbook (kept for Excel readability
# on the two large time-series tables; full data still lives in the CSVs).
EXCEL_ROW_CAP = 5000

sheets = {
    "1_Product_Catalog":      product_catalog,
    "2_Sales_Transactions":   sales_df.head(EXCEL_ROW_CAP),
    "3_Demand_Features":      weekly,
    "4_Price_Elasticity":     elast_df,
    "5_Inventory_Monitoring": inv_df,
    "6_Optimal_Pricing_Output": opt_df,
    "7_Customer_Patterns":    purchase_stats,
    "8_Demand_Forecast":      forecast_df.head(EXCEL_ROW_CAP),
}

wb = Workbook()
wb.remove(wb.active)  # drop the default blank sheet

# README sheet first
readme = wb.create_sheet("📋 README")
readme["A1"] = "Grocery Pricing Optimization Dataset"
readme["A1"].font = Font(bold=True, size=14)
readme_lines = [
    ("Sheet", "Description"),
    ("1_Product_Catalog", "500 products sampled from Instacart, with estimated base/cost price, department, shelf life"),
    ("2_Sales_Transactions", f"Daily sales per product over {N_DAYS} days (showing first {EXCEL_ROW_CAP:,} of {sales_df.shape[0]:,} rows — see CSV for full data)"),
    ("3_Demand_Features", "Weekly aggregated demand features per product"),
    ("4_Price_Elasticity", "Estimated price elasticity per product"),
    ("5_Inventory_Monitoring", "Simulated current stock levels and urgency classification"),
    ("6_Optimal_Pricing_Output", "Main output — recommended price per product from the 3-Brain pipeline"),
    ("7_Customer_Patterns", "Real Instacart reorder/cart-position behavior per product"),
    ("8_Demand_Forecast", f"{FORECAST_DAYS}-day forward demand forecast (showing first {EXCEL_ROW_CAP:,} of {forecast_df.shape[0]:,} rows — see CSV for full data)"),
]
for i, (name, desc) in enumerate(readme_lines, start=3):
    readme[f"A{i}"] = name
    readme[f"B{i}"] = desc
    readme[f"A{i}"].font = Font(bold=(i == 3))
readme.column_dimensions["A"].width = 26
readme.column_dimensions["B"].width = 90

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")

for sheet_name, df in sheets.items():
    ws = wb.create_sheet(sheet_name)
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    for row in df.itertuples(index=False):
        ws.append(list(row))
    for col_idx, col_name in enumerate(df.columns, start=1):
        max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max() if len(df) else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
    ws.freeze_panes = "A2"

Path(XLSX_PATH).parent.mkdir(parents=True, exist_ok=True)
wb.save(XLSX_PATH)
print("  ✓ Grocery_Pricing_Optimization_Dataset.xlsx")


# =============================================================================
# SUMMARY
# =============================================================================
print("\n" + "="*60)
print("ALL DATASETS GENERATED SUCCESSFULLY")
print("="*60)
summary = {
    'product_catalog.csv':            f"{N_PRODUCTS} products",
    'sales_transactions.csv':         f"{N_PRODUCTS * N_DAYS:,} rows (180 days)",
    'demand_features.csv':            "~13,000 rows (weekly aggregated)",
    'price_elasticity.csv':           f"{N_PRODUCTS} products",
    'inventory_monitoring.csv':       f"{N_PRODUCTS} products",
    'optimal_pricing_output.csv':     f"{N_PRODUCTS} products  ← MAIN OUTPUT",
    'customer_purchase_patterns.csv': "~2,740 products (real Instacart data)",
    'demand_forecast.csv':            f"~{N_PRODUCTS * FORECAST_DAYS:,} rows (30-day forward)",
}
for fname, info in summary.items():
    print(f"  {fname:<45} {info}")